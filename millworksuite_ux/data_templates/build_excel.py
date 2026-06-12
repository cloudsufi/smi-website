"""
MillworkSuite Onboarding Excel Template Builder
Generates 5 branded Excel files using the MillworkSuite colour palette.
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── Brand palette ─────────────────────────────────────────────────────────────
C_DEEP_BLACK   = "060D08"   # title bar background
C_FOREST       = "111F14"   # column header background
C_RAISED       = "1A2B1D"   # section divider background
C_GREEN        = "28C873"   # Precision Green — accent
C_GREEN_LIGHT  = "EDF8F2"   # guide row background
C_GREEN_MID    = "C8F0DA"   # example row background
C_FROST        = "E8F0EA"   # primary text on dark
C_BODY         = "2D3B30"   # body text (dark green-grey)
C_MUTED        = "6B7F6E"   # muted text
C_BORDER       = "C8D8CB"   # row borders
C_WHITE        = "FFFFFF"
C_ALT_ROW      = "F7FCF9"   # alternating data row tint
C_LOCKED       = "F2F4F2"   # read-only / system rows

# ── Reusable style factories ───────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, italic=False, size=10, color=C_BODY, name="Calibri"):
    return Font(bold=bold, italic=italic, size=size, color=color, name=name)

def border_thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def border_bottom():
    return Border(bottom=Side(style="thin", color=C_BORDER))

def wrap_center():
    return Alignment(wrap_text=True, vertical="top", horizontal="center")

def wrap_left():
    return Alignment(wrap_text=True, vertical="top", horizontal="left")

def center():
    return Alignment(vertical="center", horizontal="center")

def mid_left():
    return Alignment(vertical="center", horizontal="left")

# ── Helper: apply style to a row range ────────────────────────────────────────
def style_row(ws, row, col_start, col_end, fill_=None, font_=None, align_=None, border_=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if fill_:   cell.fill   = fill_
        if font_:   cell.font   = font_
        if align_:  cell.alignment = align_
        if border_: cell.border = border_

def write_cell(ws, row, col, value, fill_=None, font_=None, align_=None, border_=None):
    cell = ws.cell(row=row, column=col, value=value)
    if fill_:   cell.fill   = fill_
    if font_:   cell.font   = font_
    if align_:  cell.alignment = align_
    if border_: cell.border = border_
    return cell

# ── Title block (rows 1-3) ────────────────────────────────────────────────────
def write_title_block(ws, title, subtitle, purpose, num_cols):
    # Row 1 — MillworkSuite brand bar
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    c = ws.cell(row=1, column=1, value="  MillworkSuite  ·  Onboarding Data Template")
    c.fill = fill(C_DEEP_BLACK)
    c.font = font(bold=True, size=11, color=C_GREEN)
    c.alignment = mid_left()

    # Row 2 — Document title
    ws.row_dimensions[2].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c = ws.cell(row=2, column=1, value=title)
    c.fill = fill(C_FOREST)
    c.font = font(bold=True, size=14, color=C_FROST)
    c.alignment = mid_left()

    # Row 3 — Purpose / subtitle
    ws.row_dimensions[3].height = 42
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    c = ws.cell(row=3, column=1, value=purpose)
    c.fill = fill(C_RAISED)
    c.font = font(italic=True, size=9, color=C_FROST)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")

    # Row 4 — spacer
    ws.row_dimensions[4].height = 6
    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=num_cols)
    ws.cell(row=4, column=1).fill = fill(C_WHITE)

# ── Column header row ─────────────────────────────────────────────────────────
def write_col_headers(ws, row, headers, widths):
    ws.row_dimensions[row].height = 22
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = write_cell(ws, row, i, h,
                       fill_=fill(C_FOREST),
                       font_=font(bold=True, size=10, color=C_GREEN),
                       align_=Alignment(vertical="center", horizontal="left", wrap_text=True))
        c.border = border_thin()

# ── Guide row ─────────────────────────────────────────────────────────────────
def write_guide_row(ws, row, guides, num_cols):
    ws.row_dimensions[row].height = 72
    for i, g in enumerate(guides, start=1):
        c = write_cell(ws, row, i, g,
                       fill_=fill(C_GREEN_LIGHT),
                       font_=font(italic=True, size=8, color="1B7A4A"),
                       align_=wrap_left())
        c.border = border_thin()
    # label badge in col 1
    ws.cell(row=row, column=1).font = font(italic=True, bold=True, size=8, color="1B7A4A")

# ── Section divider ───────────────────────────────────────────────────────────
def write_section(ws, row, label, num_cols):
    ws.row_dimensions[row].height = 18
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=1, value=label)
    c.fill = fill(C_GREEN)
    c.font = font(bold=True, size=9, color=C_DEEP_BLACK)
    c.alignment = mid_left()

# ── Example data row ──────────────────────────────────────────────────────────
def write_example_row(ws, row, values, num_cols, locked=False):
    ws.row_dimensions[row].height = 16
    bg = C_LOCKED if locked else C_GREEN_MID
    ft = font(italic=True, size=9, color=C_MUTED) if locked else font(italic=True, size=9, color="2D6B47")
    for i in range(1, num_cols + 1):
        val = values[i-1] if i <= len(values) else ""
        c = write_cell(ws, row, i, val, fill_=fill(bg), font_=ft, align_=wrap_left())
        c.border = border_thin()
    if locked:
        ws.cell(row=row, column=1).font = font(italic=True, bold=True, size=9, color=C_MUTED)

# ── Blank data entry row ──────────────────────────────────────────────────────
def write_data_row(ws, row, num_cols, alt=False):
    ws.row_dimensions[row].height = 18
    bg = C_ALT_ROW if alt else C_WHITE
    for i in range(1, num_cols + 1):
        c = ws.cell(row=row, column=i)
        c.fill = fill(bg)
        c.font = font(size=10)
        c.alignment = wrap_left()
        c.border = border_thin()

# ── Footer ────────────────────────────────────────────────────────────────────
def write_footer(ws, row, num_cols, notes):
    ws.row_dimensions[row].height = 14
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    c = ws.cell(row=row, column=1, value=notes)
    c.fill = fill(C_DEEP_BLACK)
    c.font = font(size=8, color=C_MUTED)
    c.alignment = mid_left()

# ══════════════════════════════════════════════════════════════════════════════
# FILE 1 — USERS
# ══════════════════════════════════════════════════════════════════════════════
def build_users():
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.sheet_view.showGridLines = False

    HEADERS = ["First Name", "Last Name", "Email Address", "Role", "Department", "Notes"]
    WIDTHS  = [18, 18, 34, 24, 20, 36]
    N = len(HEADERS)

    write_title_block(ws, "User List", "",
        "PURPOSE: Every person listed here will receive an email invitation to join your MillworkSuite workspace.\n"
        "INSTRUCTIONS: Delete the EXAMPLE rows, fill in your users, and return to your MillworkSuite contact.", N)

    write_col_headers(ws, 5, HEADERS, WIDTHS)

    GUIDES = [
        "Person's first name as they should appear in MillworkSuite.",
        "Person's last name.",
        "Work email address used to sign in. Must be unique per person.",
        "Role controlling access. Must exactly match a Role Name in file 02_Roles.xlsx.\nSystem roles: Administrator | Estimator | Drafter | Viewer",
        "Optional. Team or department (e.g. Estimating / Drafting / Management / External).",
        "Optional. Special setup notes (e.g. 'start inactive', 'set up first', 'external contractor')."
    ]
    write_guide_row(ws, 6, GUIDES, N)

    examples = [
        ["Rob", "Pryor", "rob@yourcompany.com", "Administrator", "Management", "Primary admin — set up first"],
        ["Jamie", "Martinez", "jmartinez@yourcompany.com", "Estimator", "Estimating", ""],
        ["Sam", "Kowalski", "skowalski@yourcompany.com", "Drafter", "Drafting", ""],
        ["Dana", "Torres", "dtorres@yourcompany.com", "Senior Estimator", "Estimating", "Custom role — must exist in 02_Roles.xlsx"],
        ["Ellen", "Hughes", "ehughes@yourcompany.com", "Viewer", "Management", "Read-only access"],
        ["Ben", "Flores", "bflores@outsidecompany.com", "Outside Consultant", "External", "External contractor"],
    ]
    for i, ex in enumerate(examples):
        write_example_row(ws, 7 + i, ex, N)

    write_section(ws, 13, "  ↓  Enter your users below — delete the example rows above when done", N)

    for i in range(20):
        write_data_row(ws, 14 + i, N, alt=(i % 2 == 1))

    write_footer(ws, 35, N, "  MillworkSuite Onboarding Template  ·  01_Users  ·  Do not modify column headers")

    ws.freeze_panes = "A14"
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# FILE 2 — ROLES
# ══════════════════════════════════════════════════════════════════════════════
def build_roles():
    wb = Workbook()

    # ── SHEET 1: Role Definitions ──────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Role Definitions"
    ws1.sheet_view.showGridLines = False

    H1 = ["Role Name", "Type", "Description", "Assign to which users?\n(email addresses, comma-separated)", "Notes"]
    W1 = [24, 22, 48, 40, 30]
    N1 = len(H1)

    write_title_block(ws1, "Roles — Part 1: Role Definitions", "",
        "PURPOSE: Define every role that will exist in your workspace.\n"
        "System roles are pre-filled and cannot be changed. Add your own custom roles in the USER-DEFINED rows.\n"
        "In the 'Assign to which users?' column list the email addresses from 01_Users.xlsx.", N1)

    write_col_headers(ws1, 5, H1, W1)

    G1 = [
        "The exact name of the role as it will appear in MillworkSuite.\nSystem roles: leave as-is.\nUser-defined: choose a clear name (e.g. 'Senior Estimator', 'Project Manager').",
        "System = built-in, cannot be changed.\nUser-defined = created by your company's Administrator.",
        "Plain-English description of what this role allows.\nSystem roles: leave as-is.\nUser-defined: write a clear summary — this appears in the MillworkSuite roles list.",
        "List the email addresses of all users who should have this role.\nEmails must match exactly those in 01_Users.xlsx.\nSeparate multiple addresses with a comma.",
        "Optional notes for the MillworkSuite setup team."
    ]
    write_guide_row(ws1, 6, G1, N1)

    system_roles = [
        ["Administrator", "System — do not edit", "Full system access. All possible permissions.", "rob@yourcompany.com", ""],
        ["Estimator",     "System — do not edit", "Full access to projects, drawings, rooms, elevations, pricing, and AI. No drafting access.", "jmartinez@yourcompany.com", ""],
        ["Drafter",       "System — do not edit", "Project access (without delete). Drawings/Rooms/Elevations (without delete). Full drafting and AI access. No pricing.", "skowalski@yourcompany.com", ""],
        ["Viewer",        "System — do not edit", "Read-only access to project estimates. Cannot create, edit, or export.", "ehughes@yourcompany.com", ""],
    ]
    for i, r in enumerate(system_roles):
        write_example_row(ws1, 7 + i, r, N1, locked=True)

    custom_roles = [
        ["Senior Estimator",   "User-defined", "Like Estimator with added access to manage pricing and product mappings.", "dtorres@yourcompany.com", "Example — edit or delete"],
        ["Outside Consultant", "User-defined", "Limited access for external contractors — view projects and export to CAD only.", "bflores@outsidecompany.com", "Example — edit or delete"],
    ]
    for i, r in enumerate(custom_roles):
        write_example_row(ws1, 11 + i, r, N1)

    write_section(ws1, 13, "  ↓  Add your custom roles below — delete the example rows above when done", N1)
    for i in range(10):
        write_data_row(ws1, 14 + i, N1, alt=(i % 2 == 1))

    write_footer(ws1, 25, N1, "  MillworkSuite Onboarding Template  ·  02_Roles — Sheet 1  ·  System roles cannot be edited")
    ws1.freeze_panes = "A14"

    # ── SHEET 2: Permission Matrix ─────────────────────────────────────────
    ws2 = wb.create_sheet("Permission Matrix")
    ws2.sheet_view.showGridLines = False

    H2 = ["Subject", "Permission", "What this allows", "Administrator", "Estimator", "Drafter", "Viewer", "Your Custom Role\n(copy column for each)"]
    W2 = [16, 20, 44, 16, 14, 14, 14, 22]
    N2 = len(H2)

    write_title_block(ws2, "Roles — Part 2: Permission Matrix", "",
        "PURPOSE: Shows every permission available in MillworkSuite. System role columns are pre-filled for reference.\n"
        "For each USER-DEFINED role you added in Sheet 1, add a new column and enter Yes or No for each permission row.", N2)

    write_col_headers(ws2, 5, H2, W2)

    G2 = [
        "The area of the system this permission applies to.",
        "The specific action being controlled.",
        "Plain-English explanation of what this permission allows the user to do.",
        "Pre-filled. Do not change.",
        "Pre-filled. Do not change.",
        "Pre-filled. Do not change.",
        "Pre-filled. Do not change.",
        "For each custom role from Sheet 1, add a column here and enter Yes or No for every row.\nUse the exact Role Name from Sheet 1 as the column header."
    ]
    write_guide_row(ws2, 6, G2, N2)

    perms = [
        ["Projects",   "Add",                "Create a new project in the workspace",                             "Yes","Yes","Yes","No"],
        ["Projects",   "Update",             "Edit an existing project's details",                               "Yes","Yes","Yes","No"],
        ["Projects",   "Delete",             "Permanently delete a project",                                     "Yes","Yes","No", "No"],
        ["Projects",   "Import",             "Import a project definition file",                                 "Yes","Yes","No", "No"],
        ["Projects",   "Export",             "Export project data",                                              "Yes","Yes","Yes","No"],
        ["Drawings",   "Add",                "Upload new drawing sheets to a project",                          "Yes","Yes","Yes","No"],
        ["Drawings",   "Update",             "Replace or edit existing drawing sheets",                         "Yes","Yes","Yes","No"],
        ["Drawings",   "Delete",             "Remove drawing sheets from a project",                            "Yes","Yes","No", "No"],
        ["Rooms",      "Add",                "Create rooms on a drawing sheet",                                  "Yes","Yes","Yes","No"],
        ["Rooms",      "Update",             "Edit room definitions",                                            "Yes","Yes","Yes","No"],
        ["Rooms",      "Delete",             "Remove rooms from a drawing sheet",                               "Yes","Yes","No", "No"],
        ["Elevations", "Add",                "Create elevations within a room",                                  "Yes","Yes","Yes","No"],
        ["Elevations", "Update",             "Edit elevation definitions",                                       "Yes","Yes","Yes","No"],
        ["Elevations", "Delete",             "Remove elevations",                                                "Yes","Yes","No", "No"],
        ["Pricing",    "View",               "See pricing data and bid totals",                                  "Yes","Yes","No", "No"],
        ["Pricing",    "Edit",               "Change pricing on individual line items",                          "Yes","Yes","No", "No"],
        ["Pricing",    "Manage Catalogue",   "Add / edit products in the Pricing Catalogue",                    "Yes","No", "No", "No"],
        ["AI",         "Run Detection",      "Trigger AI scan on a drawing set",                                "Yes","Yes","Yes","No"],
        ["AI",         "Override Detection", "Manually correct or override an AI detection result",             "Yes","Yes","No", "No"],
        ["Drafting",   "View Drawings",      "View CAD drawings in the Drafting Catalogue",                     "Yes","No", "Yes","No"],
        ["Drafting",   "Export to CAD",      "Export a drawing to CAD software (Microvellum, Cabinet Vision…)", "Yes","No", "Yes","No"],
        ["Drafting",   "Manage Catalogue",   "Add / edit drawings in the Drafting Catalogue",                   "Yes","No", "Yes","No"],
    ]

    current_subject = None
    row = 7
    for p in perms:
        if p[0] != current_subject:
            current_subject = p[0]
            write_section(ws2, row, f"  {p[0]}", N2)
            row += 1
        ws2.row_dimensions[row].height = 16
        for col, val in enumerate(p, start=1):
            is_sys = col >= 4 and col <= 7
            bg = C_LOCKED if is_sys else C_WHITE
            ft = font(size=9, color=C_MUTED, italic=is_sys) if is_sys else font(size=9)
            c = write_cell(ws2, row, col, val, fill_=fill(bg), font_=ft, align_=wrap_left())
            c.border = border_thin()
            if col >= 4 and col <= 7:
                # color Yes/No
                if val == "Yes":
                    c.font = font(size=9, bold=True, color="1B7A4A")
                    c.fill = fill("EDF8F2")
                elif val == "No":
                    c.font = font(size=9, color=C_MUTED)
        row += 1

    write_footer(ws2, row + 1, N2, "  MillworkSuite Onboarding Template  ·  02_Roles — Sheet 2  ·  Add a column for each custom role")
    ws2.freeze_panes = "A8"
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# FILE 3 — AI CATALOGUE
# ══════════════════════════════════════════════════════════════════════════════
def build_ai_catalogue():
    wb = Workbook()
    ws = wb.active
    ws.title = "AI Catalogue"
    ws.sheet_view.showGridLines = False

    HEADERS = ["MwS AI Code", "Type Name", "Product Group", "Detection Method", "Description", "Your Notes"]
    WIDTHS  = [14, 34, 20, 18, 52, 36]
    N = len(HEADERS)

    write_title_block(ws, "AI Catalogue — Master Data (Read-Only)", "",
        "PURPOSE: This is the MillworkSuite reference list of all 21 millwork types. It is pre-filled and READ-ONLY.\n"
        "Do not change any codes, names, groups or descriptions. Use the 'Your Notes' column only to flag questions.", N)

    write_col_headers(ws, 5, HEADERS, WIDTHS)

    GUIDES = [
        "READ-ONLY. The unique MillworkSuite code for this type.\nThis code is referenced in 04_Drafting_Catalogue and 05_Pricing_Catalogue — copy it exactly.",
        "READ-ONLY. The full display name of this millwork type.",
        "READ-ONLY. The product group (Casework / Countertops / Custom Millwork / Hardware / Panels / Trim).",
        "READ-ONLY. Detectable = the AI identifies this automatically from a PDF.\nManual Entry = your estimator must enter this by hand.",
        "READ-ONLY. How detection works or how the estimator enters this type.",
        "YOUR USE ONLY. Add notes here if needed.\nExamples: 'We never use this type' · 'Confirm code maps to our SKU' · 'We call this a vanity cabinet'.\nLeave blank if no notes."
    ]
    write_guide_row(ws, 6, GUIDES, N)

    types = [
        # Detectable
        ("B1D",    "Base Cabinet — 1 Door",          "Casework",        "Detectable",    "Detected from floor plan elevation lines and single-door symbol"),
        ("B2D",    "Base Cabinet — 2 Door",          "Casework",        "Detectable",    "Detected from standard 2-door base cabinet symbol"),
        ("BSB",    "Base Cabinet — Sink Base",       "Casework",        "Detectable",    "Detected via sink cutout and no-drawer indicator"),
        ("BDB",    "Base Cabinet — Drawer Bank",     "Casework",        "Detectable",    "Detected from multi-drawer stack pattern"),
        ("W1D",    "Wall Cabinet — 1 Door",          "Casework",        "Detectable",    "Detected from wall-mount elevation symbol"),
        ("W2D",    "Wall Cabinet — 2 Door",          "Casework",        "Detectable",    "Detected from standard 2-door wall symbol"),
        ("TH",     "Tall Cabinet — Full Height",     "Casework",        "Detectable",    "Detected from full-height cabinet outline"),
        ("VB",     "Vanity Base",                    "Casework",        "Detectable",    "Detected from bathroom layout and vanity symbol"),
        ("CT-L",   "Countertop — Linear",            "Countertops",     "Detectable",    "Detected as continuous horizontal run"),
        ("CT-C",   "Countertop — Corner / Miter",   "Countertops",     "Detectable",    "Detected from L-corner junction pattern"),
        ("RD",     "Reception Desk",                 "Custom Millwork", "Detectable",    "Detected from reception desk shape and height markers"),
        ("NS",     "Nurse Station",                  "Custom Millwork", "Detectable",    "Detected from nurse station layout pattern"),
        ("WP",     "Wood Wainscot / Paneling",       "Custom Millwork", "Detectable",    "Detected from full-wall panel elevation"),
        # Manual entry
        ("CT-K",   "Countertop — Curved / ADA Knee","Countertops",     "Manual Entry",  "Non-standard curves require manual entry"),
        ("HW-P",   "Hardware — Pulls",               "Hardware",        "Manual Entry",  "Individual hardware items entered manually"),
        ("HW-H",   "Hardware — Hinges",              "Hardware",        "Manual Entry",  "Hinge counts entered per cabinet manually"),
        ("HW-DB",  "Hardware — Drawer Boxes",        "Hardware",        "Manual Entry",  "Drawer box count manually verified"),
        ("WP-FRP", "Wall Panel — FRP",               "Panels",          "Manual Entry",  "Fire-rated panel manually dimensioned"),
        ("TR-CM",  "Trim — Crown Moulding",          "Trim",            "Manual Entry",  "Moulding run entered as linear footage"),
        ("TR-BM",  "Trim — Base Moulding",           "Trim",            "Manual Entry",  "Base moulding LF entered manually"),
        ("SHS",    "Shelf Storage Unit",             "Casework",        "Manual Entry",  "Open shelving requires manual count"),
    ]

    current_section = None
    row = 7
    for t in types:
        section = "DETECTABLE — AI identifies these automatically" if t[3] == "Detectable" else "MANUAL ENTRY — estimator enters these by hand"
        if section != current_section:
            current_section = section
            write_section(ws, row, f"  {section}", N)
            row += 1

        ws.row_dimensions[row].height = 16
        for col, val in enumerate(t, start=1):
            if col == 1:  # code cell
                c = write_cell(ws, row, col, val,
                               fill_=fill(C_LOCKED),
                               font_=Font(name="Courier New", size=10, bold=True, color="1B7A4A"),
                               align_=center())
            elif col == 4:  # detection method
                det_bg = "EDF8F2" if val == "Detectable" else C_LOCKED
                det_color = "1B7A4A" if val == "Detectable" else C_MUTED
                c = write_cell(ws, row, col, val, fill_=fill(det_bg),
                               font_=font(size=9, color=det_color), align_=center())
            elif col == 6:  # notes — editable
                c = write_cell(ws, row, col, "",
                               fill_=fill(C_WHITE),
                               font_=font(size=10),
                               align_=wrap_left())
            else:
                c = write_cell(ws, row, col, val,
                               fill_=fill(C_LOCKED),
                               font_=font(size=9, color=C_BODY),
                               align_=wrap_left())
            c.border = border_thin()
        row += 1

    write_footer(ws, row + 1, N, "  MillworkSuite Onboarding Template  ·  03_AI_Catalogue  ·  Read-only — maintained by MillworkSuite")
    ws.freeze_panes = "A8"
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# FILE 4 — DRAFTING CATALOGUE
# ══════════════════════════════════════════════════════════════════════════════
def build_drafting():
    wb = Workbook()
    ws = wb.active
    ws.title = "Drafting Catalogue"
    ws.sheet_view.showGridLines = False

    HEADERS = ["Drawing Name", "File Reference ID", "MwS AI Code", "CAD Tool", "Description / Notes", "Ready to Upload?"]
    WIDTHS  = [36, 24, 16, 22, 46, 18]
    N = len(HEADERS)

    write_title_block(ws, "Drafting Catalogue", "",
        "PURPOSE: List every CAD drawing template your team uses. Each row = one drawing file.\n"
        "If you have the same type in two CAD tools (e.g. Microvellum AND Cabinet Vision) add a separate row for each.\n"
        "Valid CAD Tools: Microvellum | Cabinet Vision | Pytha | IMOS AX | Other (describe in Notes)\n"
        "Valid MwS AI Codes: see 03_AI_Catalogue.xlsx — copy the code from the MwS AI Code column.", N)

    write_col_headers(ws, 5, HEADERS, WIDTHS)

    GUIDES = [
        "The display name for this drawing in MillworkSuite.\nBe descriptive — include type, size, finish if relevant.\nExample: 'Base Cabinet 2-Door 36in' or 'Sink Base 30in Thermofoil (MV)'.",
        "Your internal file reference or drawing number.\nThis is how your team identifies the file on disk or in your file system.\nExample: MV-BC2D-36-001\nIf you have no reference number, leave blank — MillworkSuite will use the Drawing Name.",
        "The MwS AI Code this drawing corresponds to.\nCopy the code exactly from 03_AI_Catalogue.xlsx.\nExample: B2D for a 2-door base cabinet.\nOne drawing maps to one AI code.",
        "The CAD software this drawing belongs to.\nMust be one of:\nMicrovellum | Cabinet Vision | Pytha | IMOS AX | Other\nIf Other, describe in the Notes column.",
        "Material, finish, size variant, or any conditions.\nIf CAD Tool is 'Other', name the software here.",
        "Is this drawing file finalised and ready to upload?\nYes = ready\nNo = not ready yet\nNeeds Review = check with your team"
    ]
    write_guide_row(ws, 6, GUIDES, N)

    examples = [
        ["BC 2-Door 36in (Microvellum)",     "MV-BC2D-36-001",  "B2D",  "Microvellum", "Standard 2-door base 36in white shaker",     "Yes"],
        ["BC 2-Door 36in (Cabinet Vision)",  "CV-BC2D-36-001",  "B2D",  "Cabinet Vision","Same type — Cabinet Vision version",         "Yes"],
        ["Sink Base 30in",                   "MV-BSB-30-001",   "BSB",  "Microvellum", "Thermofoil sink base 30in",                  "Yes"],
        ["Solid Surface Countertop Run",     "PYTHA-CT-L-001",  "CT-L", "Pytha",       "Linear countertop run standard profile",     "Yes"],
        ["Reception Desk 2-Tier",            "IMOS-RD-2T-001",  "RD",   "IMOS AX",     "2-tier reception desk standard layout",      "Yes"],
    ]
    for i, ex in enumerate(examples):
        write_example_row(ws, 7 + i, ex, N)

    # Add data validation for CAD tool and ready columns
    dv_cad = DataValidation(type="list", formula1='"Microvellum,Cabinet Vision,Pytha,IMOS AX,Other"', allow_blank=True)
    dv_ready = DataValidation(type="list", formula1='"Yes,No,Needs Review"', allow_blank=True)
    ws.add_data_validation(dv_cad)
    ws.add_data_validation(dv_ready)

    write_section(ws, 12, "  ↓  Enter your drawings below — delete the example rows above when done", N)
    for i in range(25):
        write_data_row(ws, 13 + i, N, alt=(i % 2 == 1))
        dv_cad.add(ws.cell(row=13 + i, column=4))
        dv_ready.add(ws.cell(row=13 + i, column=6))

    write_footer(ws, 39, N, "  MillworkSuite Onboarding Template  ·  04_Drafting_Catalogue  ·  One row per drawing file per CAD tool")
    ws.freeze_panes = "A13"
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# FILE 5 — PRICING CATALOGUE
# ══════════════════════════════════════════════════════════════════════════════
def build_pricing():
    wb = Workbook()
    ws = wb.active
    ws.title = "Pricing Catalogue"
    ws.sheet_view.showGridLines = False

    HEADERS = ["Product Name", "MwS AI Code", "Pricing Method", "Unit Price ($)", "Description / Spec Notes", "Ready to Upload?"]
    WIDTHS  = [38, 16, 18, 16, 48, 18]
    N = len(HEADERS)

    write_title_block(ws, "Pricing Catalogue", "",
        "PURPOSE: List every product your company prices for millwork estimates.\n"
        "You can have multiple products mapping to the same AI type (e.g. standard and premium versions).\n"
        "Pricing Methods: Per Unit | Per Sq Ft | Per LF\n"
        "The default product used when the AI detects a type is set separately in MillworkSuite after upload.", N)

    write_col_headers(ws, 5, HEADERS, WIDTHS)

    GUIDES = [
        "The display name for this product in estimates.\nBe specific — include size, finish, and material.\nExample: 'Base Cabinet 2-Door 36in White Shaker' or 'Solid Surface Countertop — Corian'.",
        "The MwS AI Code this product corresponds to.\nCopy exactly from 03_AI_Catalogue.xlsx.\nMultiple products can share the same code (e.g. standard + premium versions of the same cabinet).",
        "How this product is priced. Choose exactly one of:\nPer Unit — flat price per detected item\nPer Sq Ft — price per square foot (you enter area per item in the estimate)\nPer LF — price per linear foot (used for countertops, trim, panels)",
        "The price in US dollars.\nFor Per Unit: price per item.\nFor Per Sq Ft: price per square foot.\nFor Per LF: price per linear foot.\nNumbers only — no $ sign or commas.\nExample: 845.00 or 28.60",
        "Material spec, finish, size range, or conditions that apply to this price.\nExample: 'White shaker, Blum hardware, standard depth'\nNote if this is a base price before add-ons.",
        "Is this product and price finalised?\nYes = ready\nNo = price TBD\nNeeds Review = check with manager"
    ]
    write_guide_row(ws, 6, GUIDES, N)

    examples = [
        ["Base Cabinet 2-Door 36in Standard",  "B2D",  "Per Unit",   "845.00",  "White shaker, Blum hardware, 36x34.5x24",           "Yes"],
        ["Base Cabinet 2-Door 36in Premium",   "B2D",  "Per Unit",   "1120.00", "Full-overlay painted premium finish",                "Yes"],
        ["Sink Base 30in Thermofoil",          "BSB",  "Per Unit",   "1210.00", "Thermofoil wrap, 30x34.5x21, no drawer bank",        "Yes"],
        ["Tall Cabinet 24in Painted",          "TH",   "Per Unit",   "1985.00", "Full height 84in, painted utility cabinet",          "Yes"],
        ["Solid Surface Countertop",           "CT-L", "Per LF",     "28.60",   "Corian or equivalent, per linear foot installed",    "Yes"],
        ["Corner Countertop Miter",            "CT-C", "Per Unit",   "385.00",  "L-corner miter joint fabrication, flat fee",         "Yes"],
        ["Reception Desk 2-Tier Standard",     "RD",   "Per Unit",   "4200.00", "2-tier, carcass only, no finish included",           "Yes"],
    ]
    for i, ex in enumerate(examples):
        write_example_row(ws, 7 + i, ex, N)

    # Data validations
    dv_method = DataValidation(type="list", formula1='"Per Unit,Per Sq Ft,Per LF"', allow_blank=True)
    dv_ready  = DataValidation(type="list", formula1='"Yes,No,Needs Review"', allow_blank=True)
    ws.add_data_validation(dv_method)
    ws.add_data_validation(dv_ready)

    write_section(ws, 14, "  ↓  Enter your products below — delete the example rows above when done", N)
    for i in range(30):
        write_data_row(ws, 15 + i, N, alt=(i % 2 == 1))
        dv_method.add(ws.cell(row=15 + i, column=3))
        dv_ready.add(ws.cell(row=15 + i, column=6))

    write_footer(ws, 46, N, "  MillworkSuite Onboarding Template  ·  05_Pricing_Catalogue  ·  Multiple products per AI code are allowed")
    ws.freeze_panes = "A15"
    return wb

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
OUT = os.path.dirname(os.path.abspath(__file__))

files = [
    ("01_Users.xlsx",              build_users()),
    ("02_Roles.xlsx",              build_roles()),
    ("03_AI_Catalogue.xlsx",       build_ai_catalogue()),
    ("04_Drafting_Catalogue.xlsx", build_drafting()),
    ("05_Pricing_Catalogue.xlsx",  build_pricing()),
]

for name, wb in files:
    path = os.path.join(OUT, name)
    wb.save(path)
    print(f"✓  {name}")

print("\nAll 5 templates saved to data_templates/")
